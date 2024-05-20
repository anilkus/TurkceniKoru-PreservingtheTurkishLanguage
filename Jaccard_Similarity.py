import pandas as pd
from sklearn.metrics import jaccard_score
from sklearn.feature_extraction.text import CountVectorizer

# Excel dosyasını oku
df = pd.read_excel("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx")  # Excel dosyasının adını ve yolunu belirtin

# Belirli hücrelerdeki cümleleri al
cumle1 = df.iloc[7, 2]  # 9. satır, 2. sütun (0'dan başlayan indeksleme)
cumle2 = df.iloc[7, 3]  # 9. satır, 3. sütun (0'dan başlayan indeksleme)

# Cümlelerin kelime kümesini oluştur
vectorizer = CountVectorizer(binary=True)
vectorizer.fit([cumle1, cumle2])
cumle1_vector = vectorizer.transform([cumle1])
cumle2_vector = vectorizer.transform([cumle2])

# Jaccard Similarity hesapla
jaccard_sim = jaccard_score(cumle1_vector.toarray()[0], cumle2_vector.toarray()[0])

# Sonucu göster
print("Jaccard Similarity:", jaccard_sim)

#excele kaydetme
'''
from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx')
ws = wb['Sheet1']


ws.cell(row=3,column=7).value = jaccard_sim



wb.save('C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx')

'''