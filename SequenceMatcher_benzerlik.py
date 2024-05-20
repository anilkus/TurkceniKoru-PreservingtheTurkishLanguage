import pandas as pd
from difflib import SequenceMatcher

def benzerlik_olcumu(cumle1, cumle2):
    benzerlik_orani = SequenceMatcher(None, cumle1, cumle2).ratio()
    return benzerlik_orani

# Excel dosyasını oku
df = pd.read_excel("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx")  # Excel dosyasının adını ve yolunu belirtin

# Belirli hücrelerdeki cümleleri al
cumle1 = df.iloc[1, 2]  # 9. satır, 2. sütun (0'dan başlayan indeksleme)
cumle2 = df.iloc[1, 3]  # 9. satır, 3. sütun (0'dan başlayan indeksleme)

# Benzerlik ölçümü
benzerlik_orani = benzerlik_olcumu(str(cumle1), str(cumle2))

# Sonucu göster
print("İki cümlenin benzerlik oranı:", benzerlik_orani)

print(cumle1)
print(cumle2)

#excele kaydetme
'''
from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx')
ws = wb['Sheet1']


ws.cell(row=3,column=5).value = benzerlik_orani



wb.save('C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx')

'''