import openai

# OpenAI API anahtarınızı buraya girin
openai.api_key = "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"


with open("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/Veri Kümesi/deneme.txt", 'r', encoding="utf-8") as text_file:
    sentence_to_check = text_file.read()
    
        

# Kullanıcıdan cümleyi alma
#sentence_to_check = input("Lütfen Türkçe olmayan kelimeleri tespit etmek için bir cümle girin: ")
#sentence_to_check = "bazen down hissedersin. bu haberlerin hepsi fake gibi geliyor. "

# Metin oluşturma örneği
prompt_text = f"tüm metinde Türkçe olmayan kelimeleri tespit et lütfen. Metin: \"{sentence_to_check}\""
response = openai.Completion.create(
  engine="gpt-3.5-turbo-instruct",  # kullanılacak model
  prompt=prompt_text,
  max_tokens=50  # oluşturulacak maksimum token sayısı
)

print(response.choices[0].text.strip())


# Metin oluşturma örneği
prompt_text = f"tespit edilen kelimeleri türkçesi ile değiştir. yabancı kelimeler hariç bir değişiklik yapma lütfen. Metin: \"{sentence_to_check}\""
response = openai.Completion.create(
  engine="gpt-3.5-turbo-instruct",  # kullanılacak model
  prompt=prompt_text,
  max_tokens=50  # oluşturulacak maksimum token sayısı
)

print(response.choices[0].text.strip())

from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/TextRank.xlsx')
ws = wb['Sayfa1']
yayın1 = "10.21673-anadoluklin.1096789-2346006 "
yayın2 = ""

recall = 0.
precision = 0.

ws.cell(row=61,column=3).value = yayın1
ws.cell(row=61,column=4).value = yayın2
ws.cell(row=61,column=5).value = gercekmetinuzunlugu
ws.cell(row=61,column=6).value = özetuzunlugu
ws.cell(row=61,column=8).value = orjinalozet
ws.cell(row=61,column=9).value = recall
ws.cell(row=61,column=10).value = precision

wb.save('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/TextRank.xlsx')
'''

'''
ozet = summary
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Çalışmanın Özetleri/TextRank/z_Anadolu Kliniği Tıp Bilimleri Dergisi/10.21673-anadoluklin.1096789-2346006.txt", "w",encoding="utf-8") as file:
    file.write(ozet)