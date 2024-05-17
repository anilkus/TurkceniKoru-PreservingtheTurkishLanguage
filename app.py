import openai

# OpenAI API key
openai.api_key = ""


with open("your text file", 'r', encoding="utf-8") as text_file:
    sentence_to_check = text_file.read()
    
        

# user input
#sentence_to_check = input("Lütfen Türkçe olmayan kelimeleri tespit etmek için bir cümle girin: ")
#sentence_to_check = "bazen down hissedersin. bu haberlerin hepsi fake gibi geliyor. "

# prompting
prompt_text = f"tüm metinde Türkçe olmayan kelimeleri tespit et lütfen. Metin: \"{sentence_to_check}\""
response = openai.Completion.create(
  engine="gpt-3.5-turbo-instruct",  # kullanılacak model
  prompt=prompt_text,
  max_tokens=50  # oluşturulacak maksimum token sayısı
)

print(response.choices[0].text.strip())


# prompt2
prompt_text = f"tespit edilen kelimeleri türkçesi ile değiştir. yabancı kelimeler hariç bir değişiklik yapma lütfen. Metin: \"{sentence_to_check}\""
response = openai.Completion.create(
  engine="gpt-3.5-turbo-instruct",  # kullanılacak model
  prompt=prompt_text,
  max_tokens=50  # oluşturulacak maksimum token sayısı
)

print(response.choices[0].text.strip())
#saving in excel 
from openpyxl import load_workbook
wb = load_workbook('excel path')
ws = wb['Sheet1']
a = your_output

ws.cell(row=3,column=3).value = your_output


wb.save('excel path')
'''

