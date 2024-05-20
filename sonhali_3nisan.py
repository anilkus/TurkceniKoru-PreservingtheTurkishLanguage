import openai

# OpenAI API anahtarınızı buraya girin
openai.api_key = "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"


with open("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/Veri Kümesi/100.txt", 'r', encoding="utf-8") as text_file:
    sentence_to_check = text_file.read()
    
        

# Kullanıcıdan cümleyi alma
#sentence_to_check = input("Lütfen Türkçe olmayan kelimeleri tespit etmek için bir cümle girin: ")
#sentence_to_check = "bazen down hissedersin. bu haberlerin hepsi fake gibi geliyor. "

# Metin oluşturma örneği
prompt_text = f"tüm metinde Türkçe olmayan kelimeleri tespit et lütfen. Metin: \"{sentence_to_check}\""
response = openai.Completion.create(
  engine="gpt-3.5-turbo-instruct",  # kullanılacak model
  prompt=prompt_text,
  max_tokens=50  # oluşturulacak maksimum token sayısı
)
a=response.choices[0].text.strip()
print(a)


# Metin oluşturma örneği
prompt_text = f"tespit edilen kelimeleri türkçesi ile değiştir. yabancı kelimeler hariç bir değişiklik yapma lütfen. Metin: \"{sentence_to_check}\""
response = openai.Completion.create(
  engine="gpt-3.5-turbo-instruct",  # kullanılacak model
  prompt=prompt_text,
  max_tokens=50  # oluşturulacak maksimum token sayısı
)
b=response.choices[0].text.strip()
print(b)


from openpyxl import load_workbook
wb = load_workbook("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/Veri Kümesi/Sonuçlar.xlsx")
ws = wb['Sheet1']
orjinalcumle = sentence_to_check
yabancikelimeler = a
uretilen= b



ws.cell(row=103,column=1).value = orjinalcumle
ws.cell(row=103,column=2).value = a
ws.cell(row=103,column=3).value = b


wb.save("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/Veri Kümesi/Sonuçlar.xlsx")


