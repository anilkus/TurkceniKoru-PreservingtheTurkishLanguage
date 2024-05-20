import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# Excel dosyasını oku
df = pd.read_excel("C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx")  # Excel dosyasının adını ve yolunu belirtin

# Belirli hücrelerdeki cümleleri al
cumle1 = df.iloc[8, 2]  # 9. satır, 2. sütun (0'dan başlayan indeksleme)
cumle2 = df.iloc[8, 3]  # 9. satır, 3. sütun (0'dan başlayan indeksleme)

# TfidfVectorizer kullanarak cümle vektörlerini oluştur
vectorizer = TfidfVectorizer()
vectorizer.fit([cumle1, cumle2])
cumle1_vector = vectorizer.transform([cumle1])
cumle2_vector = vectorizer.transform([cumle2])

# Cosine Similarity hesapla
cos_sim = cosine_similarity(cumle1_vector, cumle2_vector)

# Sonucu göster
print("Cosine Similarity:", cos_sim[0][0])


#excele kaydetme
'''
from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx')
ws = wb['Sheet1']


ws.cell(row=3,column=6).value = cos_sim[0][0]



wb.save('C:/Users/LENOVO/Desktop/Tarsus Bildiri/Yeni Bildiri/11 Mayıs/Sonuçlar.xlsx')

'''